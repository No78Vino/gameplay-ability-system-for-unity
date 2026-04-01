#!/usr/bin/env python3  
# -*- coding: utf-8 -*-  
"""  
EX-GAS ASC预设 网页编辑器 - 本地 HTTP 服务  
依赖: pip install openpyxl  
用法: python server.py --xlsx "path/to/#exgas.asc.xlsx"  
"""  
  
import argparse, json, os, re, sys, threading, webbrowser  
from http.server import BaseHTTPRequestHandler, HTTPServer  
from urllib.parse import urlparse  
  
# 将 WebEditor 根目录加入 Python 模块搜索路径  
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  
  
try:  
    import openpyxl  
except ImportError:  
    print("[ERROR] 请先运行: pip install openpyxl")  
    sys.exit(1)  
  
# ── 全局路径（由命令行参数注入）────────────────────────────────────────────────  
XLSX_PATH       = ""  
TAG_XLSX_PATH   = ""  
ATTRSET_XLSX_PATH = ""  
ABILITY_XLSX_PATH = ""  
  
# ── 常量 ────────────────────────────────────────────────────────────────────  
DATA_START_ROW = 4  
COL_ID      = 2  
COL_NAME    = 3  
COL_DESC    = 4  
COL_LEVEL   = 5  
COL_TAG     = 6  
COL_ATTRSET = 7  
COL_ABILITY = 8  
  
  
# ── Excel 读写 ────────────────────────────────────────────────────────────────  
def _parse_int_list(cell_val) -> list:  
    """容错解析整型列表，支持 ; / , / 混合文本，过滤 <=0。"""  
    if cell_val is None:  
        return []  
    s = str(cell_val).strip()  
    if not s:  
        return []  
    return [int(x) for x in re.findall(r"-?\d+", s) if int(x) > 0]  
  
  
def _join_int_list(lst: list) -> str:  
    """将 int 列表转为分号分隔的字符串"""  
    return ";".join(str(x) for x in lst)  
  
  
def read_ascs() -> list:  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.active  
    result = []  
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):  
        id_val = row[COL_ID - 1]  
        if id_val is None:  
            continue  
        try:  
            asc_id = int(id_val)  
        except (ValueError, TypeError):  
            continue  
        result.append({  
            "id":      asc_id,  
            "name":    str(row[COL_NAME    - 1] or ""),  
            "desc":    str(row[COL_DESC    - 1] or ""),  
            "level":   int(row[COL_LEVEL   - 1] or 0),  
            "tag":     _parse_int_list(row[COL_TAG     - 1]),  
            "attrSet": _parse_int_list(row[COL_ATTRSET - 1]),  
            "ability": _parse_int_list(row[COL_ABILITY - 1]),  
        })  
    wb.close()  
    return result  
  
  
def write_ascs(ascs: list):  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.active  
    # 清空旧数据区  
    for r in range(DATA_START_ROW, ws.max_row + 1):  
        for c in range(COL_ID, COL_ABILITY + 1):  
            ws.cell(row=r, column=c).value = None  
  
    row_cursor = DATA_START_ROW  
    for asc in sorted(ascs, key=lambda x: x["id"]):  
        ws.cell(row=row_cursor, column=COL_ID).value      = asc["id"]  
        ws.cell(row=row_cursor, column=COL_NAME).value    = asc["name"]  
        ws.cell(row=row_cursor, column=COL_DESC).value    = asc["desc"]  
        ws.cell(row=row_cursor, column=COL_LEVEL).value   = asc["level"]  
        ws.cell(row=row_cursor, column=COL_TAG).value     = _join_int_list(asc.get("tag", []))  
        ws.cell(row=row_cursor, column=COL_ATTRSET).value = _join_int_list(asc.get("attrSet", []))  
        ws.cell(row=row_cursor, column=COL_ABILITY).value = _join_int_list(asc.get("ability", []))  
        row_cursor += 1  
    wb.save(XLSX_PATH)  
  
  
def validate_ascs(ascs: list):  
    names = [a["name"].strip() for a in ascs]  
    if any(not n for n in names):  
        return "ASC名称不能为空"  
    if len(names) != len(set(names)):  
        return "存在重复的ASC名称"  
    ids = [a["id"] for a in ascs]  
    if len(ids) != len(set(ids)):  
        return "存在重复的ASC ID"  
    return None  
  
  
# ── HTTP Handler ──────────────────────────────────────────────────────────────  
class Handler(BaseHTTPRequestHandler):  
    def log_message(self, format, *args):  
        pass  # 静默日志  
  
    def send_json(self, data, status=200):  
        body = json.dumps(data, ensure_ascii=False).encode()  
        self.send_response(status)  
        self.send_header("Content-Type", "application/json; charset=utf-8")  
        self.send_header("Content-Length", str(len(body)))  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.end_headers()  
        self.wfile.write(body)  
  
    def send_file(self, path, mime):  
        with open(path, "rb") as f:  
            data = f.read()  
        self.send_response(200)  
        self.send_header("Content-Type", mime)  
        self.send_header("Content-Length", str(len(data)))  
        self.end_headers()  
        self.wfile.write(data)  
  
    def do_OPTIONS(self):  
        self.send_response(204)  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")  
        self.send_header("Access-Control-Allow-Headers", "Content-Type")  
        self.end_headers()  
  
    def do_GET(self):  
        p = urlparse(self.path).path  
        base = os.path.dirname(os.path.abspath(__file__))  
  
        if p in ("/", "/index.html"):  
            self.send_file(os.path.join(base, "asc_editor.html"), "text/html; charset=utf-8")  
        elif p == "/asc_editor.css":  
            self.send_file(os.path.join(base, "asc_editor.css"), "text/css; charset=utf-8")  
        elif p == "/asc_editor.js":  
            self.send_file(os.path.join(base, "asc_editor.js"), "application/javascript; charset=utf-8")  
        elif p == "/api/ascs":  
            try:  
                self.send_json({"ok": True, "ascs": read_ascs()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        elif p == "/api/info":  
            self.send_json({"ok": True, "xlsx": XLSX_PATH})  
        elif p == "/api/choices/tags":  
            try:  
                from gas_xlsx_choice import GasXlsxChoice  
                c = GasXlsxChoice({"tag": TAG_XLSX_PATH})  
                self.send_json({"ok": True, "tags": c.tags()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        elif p == "/api/choices/attrsets":  
            try:  
                from gas_xlsx_choice import GasXlsxChoice  
                c = GasXlsxChoice({"attrset": ATTRSET_XLSX_PATH})  
                self.send_json({"ok": True, "attrsets": c.attrsets()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        elif p == "/api/choices/abilities":  
            try:  
                from gas_xlsx_choice import GasXlsxChoice  
                c = GasXlsxChoice({"ability": ABILITY_XLSX_PATH})  
                self.send_json({"ok": True, "abilities": c.abilities()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def read_body(self):  
        length = int(self.headers.get("Content-Length", 0))  
        return json.loads(self.rfile.read(length)) if length else {}  
  
    def do_POST(self):  
        p = urlparse(self.path).path  
        if p == "/api/ascs":  
            try:  
                data = self.read_body()  
                ascs = read_ascs()  
                new_id = data.get("id")  
                if not new_id:  
                    new_id = max((a["id"] for a in ascs), default=0) + 1  
                new_asc = {  
                    "id":      int(new_id),  
                    "name":    data.get("name", "").strip(),  
                    "desc":    data.get("desc", ""),  
                    "level":   int(data.get("level", 1)),  
                    "tag":     data.get("tag", []),  
                    "attrSet": data.get("attrSet", []),  
                    "ability": data.get("ability", []),  
                }  
                ascs.append(new_asc)  
                err = validate_ascs(ascs)  
                if err:  
                    self.send_json({"ok": False, "error": err}, 400)  
                    return  
                write_ascs(ascs)  
                self.send_json({"ok": True, "asc": new_asc})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_PUT(self):  
        p = urlparse(self.path).path  
        if p.startswith("/api/ascs/"):  
            try:  
                old_id = int(p.split("/")[-1])  
                data   = self.read_body()  
                ascs   = read_ascs()  
                idx = next((i for i, a in enumerate(ascs) if a["id"] == old_id), None)  
                if idx is None:  
                    self.send_json({"ok": False, "error": "未找到该ASC预设"}, 404)  
                    return  
                ascs[idx] = {  
                    "id":      int(data.get("id", old_id)),  
                    "name":    data.get("name", "").strip(),  
                    "desc":    data.get("desc", ""),  
                    "level":   int(data.get("level", 1)),  
                    "tag":     data.get("tag", []),  
                    "attrSet": data.get("attrSet", []),  
                    "ability": data.get("ability", []),  
                }  
                err = validate_ascs(ascs)  
                if err:  
                    self.send_json({"ok": False, "error": err}, 400)  
                    return  
                write_ascs(ascs)  
                self.send_json({"ok": True, "asc": ascs[idx]})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_DELETE(self):  
        p = urlparse(self.path).path  
        if p.startswith("/api/ascs/"):  
            try:  
                del_id = int(p.split("/")[-1])  
                ascs   = read_ascs()  
                ascs   = [a for a in ascs if a["id"] != del_id]  
                write_ascs(ascs)  
                self.send_json({"ok": True})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
            
# ── 入口 ──────────────────────────────────────────────────────────────────────  
def main():  
    global XLSX_PATH, TAG_XLSX_PATH, ATTRSET_XLSX_PATH, ABILITY_XLSX_PATH  
    ap = argparse.ArgumentParser(description="EX-GAS ASC预设 网页编辑器服务")  
    ap.add_argument("--xlsx",          required=True, help="#exgas.asc.xlsx 路径")  
    ap.add_argument("--port",          type=int, default=8768)  
    ap.add_argument("--no-browser",    action="store_true")  
    ap.add_argument("--tag-xlsx",      default="", help="#exgas.gameplayTags.xlsx 路径")  
    ap.add_argument("--attrset-xlsx",  default="", help="#exgas.attributeSet.xlsx 路径")  
    ap.add_argument("--ability-xlsx",  default="", help="#exgas.ability.xlsx 路径")  
    args = ap.parse_args()  
  
    XLSX_PATH         = os.path.abspath(args.xlsx)  
    TAG_XLSX_PATH     = os.path.abspath(args.tag_xlsx)     if args.tag_xlsx     else ""  
    ATTRSET_XLSX_PATH = os.path.abspath(args.attrset_xlsx) if args.attrset_xlsx else ""  
    ABILITY_XLSX_PATH = os.path.abspath(args.ability_xlsx) if args.ability_xlsx else ""  
  
    if not os.path.exists(XLSX_PATH):  
        print(f"[ERROR] 文件不存在: {XLSX_PATH}")  
        sys.exit(1)  
  
    url = f"http://127.0.0.1:{args.port}"  
    print(f"[EX-GAS ASC Editor] {url}")  
    print(f"  Excel: {XLSX_PATH}")  
    print(f"  Ctrl+C 停止")  
  
    if not args.no_browser:  
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()  
  
    try:  
        HTTPServer(("127.0.0.1", args.port), Handler).serve_forever()  
    except KeyboardInterrupt:  
        print("\n[停止]")  
  
  
if __name__ == "__main__":  
    main()
