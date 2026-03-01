#!/usr/bin/env python3  
# -*- coding: utf-8 -*-  
"""  
EX-GAS GameplayTag 网页编辑器 - 本地 HTTP 服务  
依赖: pip install openpyxl  
用法: python server.py --xlsx "path/to/#exgas.gameplayTags.xlsx"  
"""  
  
import argparse, json, os, sys, threading, webbrowser  
from http.server import BaseHTTPRequestHandler, HTTPServer  
from urllib.parse import urlparse  
  
try:  
    import openpyxl  
except ImportError:  
    print("[ERROR] 请先运行: pip install openpyxl")  
    sys.exit(1)  
  
# ── 常量 ────────────────────────────────────────────────────────────────────  
DATA_START_ROW = 4   # 数据从第4行开始（1-3行为表头+Luban类型定义）  
COL_ID   = 2         # 第2列: ID  
COL_NAME = 3         # 第3列: Name  
COL_DESC = 4         # 第4列: Desc  
  
XLSX_PATH = ""       # 由命令行参数注入  
  
STATIC_TYPES = {  
    ".html": "text/html; charset=utf-8",  
    ".css":  "text/css; charset=utf-8",  
    ".js":   "application/javascript; charset=utf-8",  
}  
  
# ── Excel 读写层 ─────────────────────────────────────────────────────────────  
  
def read_tags():  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.worksheets[0]  
    tags = []  
    row = DATA_START_ROW  
    while ws.cell(row=row, column=COL_ID).value is not None:  
        tags.append({  
            "id":   int(ws.cell(row=row, column=COL_ID).value),  
            "name": str(ws.cell(row=row, column=COL_NAME).value or ""),  
            "desc": str(ws.cell(row=row, column=COL_DESC).value or ""),  
        })  
        row += 1  
    wb.close()  
    return tags  
  
def write_tags(tags):  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.worksheets[0]  
    # 清空旧数据（保留第1-3行：表头+Luban类型行）  
    for r in range(DATA_START_ROW, ws.max_row + 1):  
        ws.cell(row=r, column=COL_ID).value   = None  
        ws.cell(row=r, column=COL_NAME).value = None  
        ws.cell(row=r, column=COL_DESC).value = None  
    # 写入新数据（按ID排序）  
    for i, tag in enumerate(sorted(tags, key=lambda t: t["id"])):  
        r = DATA_START_ROW + i  
        ws.cell(row=r, column=COL_ID).value   = tag["id"]  
        ws.cell(row=r, column=COL_NAME).value = tag["name"]  
        ws.cell(row=r, column=COL_DESC).value = tag["desc"]  
    wb.save(XLSX_PATH)  
    wb.close()  
  
def next_id(tags):  
    return max((t["id"] for t in tags), default=1000) + 1

def validate_tags(tags):
    """校验：name不为空、不重复、ID不重复。返回 error string 或 None。"""
    names = [t["name"].strip() for t in tags]
    if any(not n for n in names):
        return "Tag名称不能为空"
    if len(names) != len(set(names)):
        return "存在重复的Tag名称"
    ids = [t["id"] for t in tags]
    if len(ids) != len(set(ids)):
        return "存在重复的Tag ID"
    return None

def get_warnings(tags):
    """返回警告列表（不阻止保存）：缺少父类的Tag。"""
    name_set = {t["name"] for t in tags}
    warnings = []
    for t in tags:
        parts = t["name"].split(".")
        for i in range(1, len(parts)):
            parent = ".".join(parts[:i])
            if parent not in name_set:
                warnings.append(f"Tag「{t['name']}」缺少父类「{parent}」")
    return warnings
  
# ── HTTP 处理器 ──────────────────────────────────────────────────────────────  
  
class Handler(BaseHTTPRequestHandler):  
  
    def log_message(self, fmt, *args):  
        print(f"  [{args[1]}] {self.command} {self.path}")  
  
    def send_json(self, data, status=200):  
        body = json.dumps(data, ensure_ascii=False).encode()  
        self.send_response(status)  
        self.send_header("Content-Type", "application/json; charset=utf-8")  
        self.send_header("Content-Length", len(body))  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.end_headers()  
        self.wfile.write(body)  
  
    def read_json(self):  
        n = int(self.headers.get("Content-Length", 0))  
        return json.loads(self.rfile.read(n))  
  
    def do_OPTIONS(self):  
        self.send_response(204)  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,DELETE,OPTIONS")  
        self.send_header("Access-Control-Allow-Headers", "Content-Type")  
        self.end_headers()  
  
    def do_GET(self):  
        p = urlparse(self.path).path  
        if p == "/":  
            p = "/tag_editor.html"  
  
        # 静态文件服务（.html / .css / .js）  
        ext = os.path.splitext(p)[1]  
        if ext in STATIC_TYPES:  
            filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), p.lstrip("/"))  
            if os.path.exists(filepath):  
                body = open(filepath, "rb").read()  
                self.send_response(200)  
                self.send_header("Content-Type", STATIC_TYPES[ext])  
                self.send_header("Content-Length", len(body))  
                self.end_headers()  
                self.wfile.write(body)  
            else:  
                self.send_json({"ok": False, "error": f"File not found: {p}"}, 404)  
            return  
  
        # API 路由  
        if p == "/api/tags":  
            try:  
                self.send_json({"ok": True, "tags": read_tags()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        elif p == "/api/info":  
            self.send_json({"ok": True, "xlsx": XLSX_PATH})  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_POST(self):  
        p = urlparse(self.path).path  
        if p == "/api/tags":  
            try:  
                body = self.read_json()  
                tags = read_tags()  
                custom_id = body.get("id")  
                if custom_id is not None:  
                    custom_id = int(custom_id)  
                    if any(t["id"] == custom_id for t in tags):  
                        return self.send_json({"ok": False, "error": f"ID {custom_id} 已存在"}, 400)  
                    new_id = custom_id  
                else:  
                    new_id = next_id(tags)  
                new_tag = {  
                    "id":   new_id,  
                    "name": body.get("name", "").strip(),  
                    "desc": body.get("desc", "").strip(),  
                }  
                err = validate_tags(tags + [new_tag])  
                if err:  
                    return self.send_json({"ok": False, "error": err}, 400)  
                write_tags(tags + [new_tag])  
                self.send_json({"ok": True, "tag": new_tag})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)

    def do_PUT(self):
        parts = urlparse(self.path).path.strip("/").split("/")
        if len(parts) == 3 and parts[:2] == ["api", "tags"]:
            try:
                tid = int(parts[2])
                body = self.read_json()
                tags = read_tags()
                tag = next((t for t in tags if t["id"] == tid), None)
                if not tag:
                    return self.send_json({"ok": False, "error": f"ID不存在: {tid}"}, 404)
                    # 支持修改 ID  
                new_id = body.get("id")
                if new_id is not None:
                    new_id = int(new_id)
                    if new_id != tid and any(t["id"] == new_id for t in tags):
                        return self.send_json({"ok": False, "error": f"ID {new_id} 已存在"}, 400)
                    tag["id"] = new_id
                tag["name"] = body.get("name", tag["name"]).strip()
                tag["desc"] = body.get("desc", tag["desc"]).strip()
                err = validate_tags(tags)
                if err:
                    return self.send_json({"ok": False, "error": err}, 400)
                write_tags(tags)
                warnings = get_warnings(tags)
                self.send_json({"ok": True, "warnings": warnings})
            except Exception as e:
                self.send_json({"ok": False, "error": str(e)}, 500)
  
    def do_DELETE(self):  
        # DELETE /api/tags/{id}  
        parts = urlparse(self.path).path.strip("/").split("/")  
        if len(parts) == 3 and parts[:2] == ["api", "tags"]:  
            try:  
                tid = int(parts[2])  
                tags = read_tags()  
                new_tags = [t for t in tags if t["id"] != tid]  
                if len(new_tags) == len(tags):  
                    return self.send_json({"ok": False, "error": f"ID不存在: {tid}"}, 404)  
                write_tags(new_tags)  
                self.send_json({"ok": True})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
# ── 入口 ────────────────────────────────────────────────────────────────────  
  
def main():  
    global XLSX_PATH  
    ap = argparse.ArgumentParser(description="EX-GAS GameplayTag 网页编辑器服务")  
    ap.add_argument("--xlsx", required=True, help="#exgas.gameplayTags.xlsx 路径")  
    ap.add_argument("--port", type=int, default=8765)  
    ap.add_argument("--no-browser", action="store_true")  
    args = ap.parse_args()  
  
    XLSX_PATH = os.path.abspath(args.xlsx)  
    if not os.path.exists(XLSX_PATH):  
        print(f"[ERROR] 文件不存在: {XLSX_PATH}")  
        sys.exit(1)  
  
    url = f"http://127.0.0.1:{args.port}"  
    print(f"[EX-GAS Tag Editor] {url}")  
    print(f"  Excel: {XLSX_PATH}")  
    print(f"  Ctrl+C 停止")  
  
    if not args.no_browser:  
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()  
  
    try:  
        HTTPServer(("127.0.0.1", args.port), Handler).serve_forever()  
    except KeyboardInterrupt:  
        print("\n[EX-GAS Tag Editor] 服务已停止")  
  
if __name__ == "__main__":  
    main()