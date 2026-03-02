#!/usr/bin/env python3  
# -*- coding: utf-8 -*-  
"""  
EX-GAS AttributeSet 网页编辑器 - 本地 HTTP 服务  
依赖: pip install openpyxl  
用法: python server.py --xlsx "path/to/#exgas.attributeSet.xlsx"  
"""  


import argparse, json, os, sys, threading, webbrowser  
from http.server import BaseHTTPRequestHandler, HTTPServer  
from urllib.parse import urlparse, parse_qs  

# 将 WebEditor 根目录加入 Python 模块搜索路径  
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
 
try:  
    import openpyxl  
except ImportError:  
    print("[ERROR] 请先运行: pip install openpyxl")  
    sys.exit(1)  


ATTR_XLSX_PATH = ""   # 由 --attr-xlsx 参数注入

# ── 常量 ────────────────────────────────────────────────────────────────────  
DATA_START_ROW  = 5  
COL_SET_ID      = 2   # 第2列: AttributeSet ID  
COL_SET_NAME    = 3   # 第3列: Name  
COL_SET_DESC    = 4   # 第4列: Desc  
COL_ATTR_ID     = 5   # 第5列: Attribute.ID  
# 第6列: 注释列（不导出，跳过）  
COL_ATTR_INIT   = 7   # 第7列: Attribute.InitValue  
COL_ATTR_MIN    = 8   # 第8列: Attribute.MinValue  
COL_ATTR_MAX    = 9   # 第9列: Attribute.MaxValue  
COL_ATTR_USEMIN = 10  # 第10列: Attribute.UseMinValue  
COL_ATTR_USEMAX = 11  # 第11列: Attribute.UseMaxValue   
  
XLSX_PATH = ""       # 由命令行参数注入  
  
  
# ── Excel 读写 ───────────────────────────────────────────────────────────────  
def read_attrsets():  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.active  
    result = []  
    current_set = None  
  
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):  
        set_id_val  = row[COL_SET_ID - 1]  
        attr_id_val = row[COL_ATTR_ID - 1]  
  
        if set_id_val is not None:  
            # 新的 AttributeSet 主行  
            try:  
                set_id_int = int(set_id_val)  
            except (ValueError, TypeError):  
                continue  
            current_set = {  
                "id":         set_id_int,  
                "name":       str(row[COL_SET_NAME - 1] or ""),  
                "desc":       str(row[COL_SET_DESC - 1] or ""),  
                "attributes": [],  
            }  
            result.append(current_set)  
  
        # 同一行或续行：解析 Attribute 条目  
        if attr_id_val is not None and current_set is not None:  
            try:  
                current_set["attributes"].append({  
                    "id":         int(attr_id_val),  
                    "initValue":  float(row[COL_ATTR_INIT - 1] or 0),  
                    "minValue":   float(row[COL_ATTR_MIN - 1] or 0),  
                    "maxValue":   float(row[COL_ATTR_MAX - 1] or 0),  
                    "useMinValue": str(row[COL_ATTR_USEMIN - 1] or "").strip().lower() in ("true", "1"),  
                    "useMaxValue": str(row[COL_ATTR_USEMAX - 1] or "").strip().lower() in ("true", "1"),  
                })  
            except (ValueError, TypeError):  
                pass  
        # 两列都为 None：空行，跳过  
  
    wb.close()  
    return result  
  
  
def write_attrsets(attrsets):  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.active  
  
    # 清空旧数据区  
    for r in range(DATA_START_ROW, ws.max_row + 1):  
        for c in range(COL_SET_ID, COL_ATTR_USEMAX + 1):  
            ws.cell(row=r, column=c).value = None  
  
    row_cursor = DATA_START_ROW  
    for attrset in sorted(attrsets, key=lambda s: s["id"]):  
        attrs = attrset.get("attributes", [])  
        row_count = max(len(attrs), 1)  
        for i in range(row_count):  
            r = row_cursor + i  
            if i == 0:  
                ws.cell(row=r, column=COL_SET_ID).value   = attrset["id"]  
                ws.cell(row=r, column=COL_SET_NAME).value = attrset["name"]  
                ws.cell(row=r, column=COL_SET_DESC).value = attrset["desc"]  
            if i < len(attrs):  
                a = attrs[i]  
                ws.cell(row=r, column=COL_ATTR_ID).value     = a["id"]  
                ws.cell(row=r, column=COL_ATTR_INIT).value   = a["initValue"]  
                ws.cell(row=r, column=COL_ATTR_MIN).value    = a["minValue"]  
                ws.cell(row=r, column=COL_ATTR_MAX).value    = a["maxValue"]  
                ws.cell(row=r, column=COL_ATTR_USEMIN).value = "true" if a.get("useMinValue") else "false"  
                ws.cell(row=r, column=COL_ATTR_USEMAX).value = "true" if a.get("useMaxValue") else "false"  
        row_cursor += row_count  
  
    wb.save(XLSX_PATH)
  
  
# ── 校验 ─────────────────────────────────────────────────────────────────────  
def validate_attrsets(attrsets):  
    """校验：name不为空、不重复、ID不重复。返回 error string 或 None。"""  
    names = [s["name"].strip() for s in attrsets]  
    if any(not n for n in names):  
        return "属性集名称不能为空"  
    if len(names) != len(set(names)):  
        return "存在重复的属性集名称"  
    ids = [s["id"] for s in attrsets]  
    if len(ids) != len(set(ids)):  
        return "存在重复的属性集 ID"  
    # 校验每个 AttributeSet 内部的 Attribute ID 不重复  
    for s in attrsets:  
        attr_ids = [a["id"] for a in s.get("attributes", [])]  
        if len(attr_ids) != len(set(attr_ids)):  
            return f"属性集「{s['name']}」内存在重复的 Attribute ID"  
    return None  
  
  
# ── HTTP Handler ─────────────────────────────────────────────────────────────  
class Handler(BaseHTTPRequestHandler):  
    def log_message(self, format, *args):  
        pass  # 静默日志  
  
    def send_json(self, data, status=200):  
        body = json.dumps(data, ensure_ascii=False).encode()  
        self.send_response(status)  
        self.send_header("Content-Type", "application/json; charset=utf-8")  
        self.send_header("Content-Length", str(len(body)))  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.end_headers()  
        self.wfile.write(body)  
  
    def send_file(self, path, mime):  
        with open(path, "rb") as f:  
            data = f.read()  
        self.send_response(200)  
        self.send_header("Content-Type", mime)  
        self.send_header("Content-Length", str(len(data)))  
        self.end_headers()  
        self.wfile.write(data)  
  
    def do_OPTIONS(self):  
        self.send_response(204)  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")  
        self.send_header("Access-Control-Allow-Headers", "Content-Type")  
        self.end_headers()  
  
    def do_GET(self):  
        p = urlparse(self.path).path  
        base = os.path.dirname(os.path.abspath(__file__))  
  
        if p == "/" or p == "/index.html":  
            self.send_file(os.path.join(base, "attrset_editor.html"), "text/html; charset=utf-8")  
        elif p == "/attrset_editor.css":  
            self.send_file(os.path.join(base, "attrset_editor.css"), "text/css; charset=utf-8")  
        elif p == "/attrset_editor.js":  
            self.send_file(os.path.join(base, "attrset_editor.js"), "application/javascript; charset=utf-8")  
        elif p == "/api/attrsets":  
            try:  
                self.send_json({"ok": True, "attrsets": read_attrsets()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        elif p == "/api/info":  
            self.send_json({"ok": True, "xlsx": XLSX_PATH})   
        elif p == "/api/choices/attrs":  
            try:  
                from gas_xlsx_choice import GasXlsxChoice  
                c = GasXlsxChoice({"attr": ATTR_XLSX_PATH})  
                self.send_json({"ok": True, "attrs": c.attrs()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def read_body(self):  
        length = int(self.headers.get("Content-Length", 0))  
        return json.loads(self.rfile.read(length)) if length else {}  
  
    def do_POST(self):  
        p = urlparse(self.path).path  
        if p == "/api/attrsets":  
            try:  
                data    = self.read_body()  
                attrsets = read_attrsets()  
                # 自动分配 ID  
                new_id = data.get("id")  
                if not new_id:  
                    new_id = max((s["id"] for s in attrsets), default=0) + 1  
                new_set = {  
                    "id":         int(new_id),  
                    "name":       data.get("name", "").strip(),  
                    "desc":       data.get("desc", ""),  
                    "attributes": data.get("attributes", []),  
                }  
                attrsets.append(new_set)  
                err = validate_attrsets(attrsets)  
                if err:  
                    self.send_json({"ok": False, "error": err}, 400)  
                    return  
                write_attrsets(attrsets)  
                self.send_json({"ok": True, "attrset": new_set})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_PUT(self):  
        p = urlparse(self.path).path  
        # /api/attrsets/{id}  
        if p.startswith("/api/attrsets/"):  
            try:  
                old_id   = int(p.split("/")[-1])  
                data     = self.read_body()  
                attrsets = read_attrsets()  
                idx = next((i for i, s in enumerate(attrsets) if s["id"] == old_id), None)  
                if idx is None:  
                    self.send_json({"ok": False, "error": "未找到该属性集"}, 404)  
                    return  
                attrsets[idx] = {  
                    "id":         int(data.get("id", old_id)),  
                    "name":       data.get("name", "").strip(),  
                    "desc":       data.get("desc", ""),  
                    "attributes": data.get("attributes", []),  
                }  
                err = validate_attrsets(attrsets)  
                if err:  
                    self.send_json({"ok": False, "error": err}, 400)  
                    return  
                write_attrsets(attrsets)  
                self.send_json({"ok": True, "attrset": attrsets[idx]})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_DELETE(self):  
        p = urlparse(self.path).path  
        if p.startswith("/api/attrsets/"):  
            try:  
                del_id   = int(p.split("/")[-1])  
                attrsets = read_attrsets()  
                attrsets = [s for s in attrsets if s["id"] != del_id]  
                write_attrsets(attrsets)  
                self.send_json({"ok": True})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
  
# ── 入口 ─────────────────────────────────────────────────────────────────────  
def main():  
    global XLSX_PATH  
    ap = argparse.ArgumentParser(description="EX-GAS AttributeSet 网页编辑器服务")  
    ap.add_argument("--xlsx", required=True, help="#exgas.attributeSet.xlsx 路径")  
    ap.add_argument("--port", type=int, default=8767)  
    ap.add_argument("--no-browser", action="store_true")  
    ap.add_argument("--attr-xlsx", default="", help="#exgas.attribute.xlsx 路径（用于属性下拉选择）")
    args = ap.parse_args()  
  
    global ATTR_XLSX_PATH  
    ATTR_XLSX_PATH = os.path.abspath(args.attr_xlsx) if args.attr_xlsx else ""
  
    XLSX_PATH = os.path.abspath(args.xlsx)  
    if not os.path.exists(XLSX_PATH):  
        print(f"[ERROR] 文件不存在: {XLSX_PATH}")  
        sys.exit(1)  
  
    url = f"http://127.0.0.1:{args.port}"  
    print(f"[EX-GAS AttributeSet Editor] {url}")  
    print(f"  Excel: {XLSX_PATH}")  
    print(f"  Ctrl+C 停止")  
  
    if not args.no_browser:  
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()  
  
    try:  
        HTTPServer(("127.0.0.1", args.port), Handler).serve_forever()  
    except KeyboardInterrupt:  
        print("\n[停止]")  
  
  
if __name__ == "__main__":  
    main()