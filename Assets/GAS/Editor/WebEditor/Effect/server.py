#!/usr/bin/env python3  
# -*- coding: utf-8 -*-  
"""  
EX-GAS GameplayEffect 网页编辑器 - 本地 HTTP 服务  
依赖: pip install openpyxl  
用法: python server.py --xlsx "path/to/#exgas.gameplayEffects.xlsx"  
"""  
  
import argparse, json, os, re, sys, threading, webbrowser  
from http.server import BaseHTTPRequestHandler, HTTPServer  
from urllib.parse import urlparse, parse_qs  
  
try:  
    import openpyxl  
except ImportError:  
    print("[ERROR] 请先运行: pip install openpyxl")  
    sys.exit(1)  
  
# ── 常量 ────────────────────────────────────────────────────────────────────  
DATA_START_ROW = 4   # 数据从第4行开始（1-3行为表头+Luban类型定义）  
  
COL_MAP = {}  
  
XLSX_PATH = ""  
TAG_XLSX_PATH = ""  
ATTRSET_XLSX_PATH = ""  
ABILITY_XLSX_PATH = ""  
CUE_XLSX_PATH = ""  
MMC_XLSX_PATH = ""
ATTR_XLSX_PATH = ""

STATIC_TYPES = {  
    ".html": "text/html; charset=utf-8",  
    ".css":  "text/css; charset=utf-8",  
    ".js":   "application/javascript; charset=utf-8",  
}  
  
# ── 枚举映射 ────────────────────────────────────────────────────────────────  
ENUM_TIME_UNIT = ["Frame", "Turn"]  
ENUM_OPERATION = {"Add": 0, "Multiply": 1, "Override": 2, "Minus": 3, "Divide": 4}  
ENUM_OPERATION_REV = {0: "Add", 1: "Multiply", 2: "Override", 3: "Minus", 4: "Divide"}  
ENUM_STACKING_TYPE = ["AggregateBySource", "AggregateByTarget"]  
ENUM_DURATION_REFRESH = ["NeverRefresh", "RefreshOnSuccessfulApplication"]  
ENUM_PERIOD_RESET = ["NeverRefresh", "ResetOnSuccessfulApplication"]  
ENUM_EXPIRATION = ["ClearEntireStack", "RemoveSingleStackAndRefreshDuration", "RefreshDuration"]  
ENUM_ABILITY_ACTIVATE = ["None", "WhenAdded", "SyncWithEffect"]  
ENUM_ABILITY_DEACTIVATE = ["None", "SyncWithEffect"]  
ENUM_ABILITY_REMOVE = ["None", "SyncWithEffect", "WhenEnd", "WhenCancel", "WhenCancelOrEnd"]  

TAG_REQUIREMENT_PROTOCOL_FIELDS = [
    {"component": "ApplicationRequiredTags", "excel_header": "ApplicationRequiredTags", "json_key": "applicationRequiredTags", "mode": "all"},
    {"component": "OngoingRequiredTags", "excel_header": "OngoingRequiredTags", "json_key": "ongoingRequiredTags", "mode": "all"},
    {"component": "RemoveGameplayEffectsWithTags", "excel_header": "RemoveGameplayEffectsWithTags", "json_key": "removeEffectsWithTags", "mode": "any"},
    {"component": "ImmunityTags", "excel_header": "ImmunityTags", "json_key": "immunityTags", "mode": "any"},
]
  
# ── 辅助数据读取 ─────────────────────────────────────────────────────────────  
  
def read_tags_for_dropdown():  
    """读取Tag数据用于下拉选择"""  
    if not TAG_XLSX_PATH or not os.path.exists(TAG_XLSX_PATH):  
        return []  
    try:  
        wb = openpyxl.load_workbook(TAG_XLSX_PATH)  
        ws = wb.worksheets[0]  
        tags = []  
        row = DATA_START_ROW  
        while ws.cell(row=row, column=2).value is not None:  
            tag_id = ws.cell(row=row, column=2).value  
            tag_name = ws.cell(row=row, column=3).value  
            if tag_id is not None:  
                tags.append({"id": int(tag_id), "name": str(tag_name or "")})  
            row += 1  
        wb.close()  
        return tags  
    except Exception as e:  
        print(f"[WARN] 读取Tag失败: {e}")  
        return []  
  
def read_abilities_for_dropdown():  
    """读取Ability数据用于下拉选择"""  
    if not ABILITY_XLSX_PATH or not os.path.exists(ABILITY_XLSX_PATH):  
        return []  
    try:  
        wb = openpyxl.load_workbook(ABILITY_XLSX_PATH)  
        ws = wb.worksheets[0]  
        abilities = []  
        row = DATA_START_ROW  
        while ws.cell(row=row, column=2).value is not None:  
            ability_id = ws.cell(row=row, column=2).value  
            ability_name = ws.cell(row=row, column=3).value  
            if ability_id is not None:  
                abilities.append({"id": int(ability_id), "name": str(ability_name or "")})  
            row += 1  
        wb.close()  
        return abilities  
    except Exception as e:  
        print(f"[WARN] 读取Ability失败: {e}")  
        return []  
  
def read_cues_for_dropdown():  
    """读取Cue数据用于下拉选择"""  
    if not CUE_XLSX_PATH or not os.path.exists(CUE_XLSX_PATH):  
        return []  
    try:  
        wb = openpyxl.load_workbook(CUE_XLSX_PATH)  
        ws = wb.worksheets[0]  
        cues = []  
        row = DATA_START_ROW  
        while ws.cell(row=row, column=2).value is not None:  
            cue_id = ws.cell(row=row, column=2).value  
            cue_name = ws.cell(row=row, column=3).value  
            if cue_id is not None:  
                cues.append({"id": int(cue_id), "name": str(cue_name or "")})  
            row += 1  
        wb.close()  
        return cues  
    except Exception as e:  
        print(f"[WARN] 读取Cue失败: {e}")  
        return []  

def read_mmcs_for_dropdown():  
    """读取MMC数据用于下拉选择"""  
    if not MMC_XLSX_PATH or not os.path.exists(MMC_XLSX_PATH):  
        return []  
    try:  
        wb = openpyxl.load_workbook(MMC_XLSX_PATH)  
        ws = wb.worksheets[0]  
        mmcs = []  
        row = DATA_START_ROW  
        while ws.cell(row=row, column=2).value is not None:  
            mmc_id = ws.cell(row=row, column=2).value  
            mmc_name = ws.cell(row=row, column=3).value  
            if mmc_id is not None:  
                mmcs.append({"id": int(mmc_id), "name": str(mmc_name or "")})  
            row += 1  
        wb.close()  
        return mmcs  
    except Exception as e:  
        print(f"[WARN] 读取MMC失败: {e}")  
        return []
          
def read_attrsets_for_dropdown():  
    """读取AttributeSet数据用于下拉选择（Modifier 的 AttrSet/Attr 级联下拉）  
  
    修复：  
    1. AttributeSet xlsx 数据从第5行开始（不是第4行）  
    2. 列结构为多行格式：col2=SetID, col3=Name, col4=Desc, col5=Attribute.ID  
       主行 col2 有值，子行 col2 为 None  
    3. 需要 attr.xlsx 做 id->name 映射  
    """  
    ATTRSET_DATA_START_ROW = 5  # AttributeSet 表从第5行开始  
  
    if not ATTRSET_XLSX_PATH or not os.path.exists(ATTRSET_XLSX_PATH):  
        return []  
    try:  
        # ── 步骤1：构建 attr id -> name 映射 ──  
        attr_name_map = {}  
        if ATTR_XLSX_PATH and os.path.exists(ATTR_XLSX_PATH):  
            try:  
                wb_attr = openpyxl.load_workbook(ATTR_XLSX_PATH, read_only=True)  
                ws_attr = wb_attr.worksheets[0]  
                row_a = 4  # attr.xlsx 从第4行开始  
                while True:  
                    aid = ws_attr.cell(row=row_a, column=2).value  
                    if aid is None:  
                        break  
                    aname = ws_attr.cell(row=row_a, column=3).value  
                    try:  
                        attr_name_map[int(aid)] = str(aname or "")  
                    except (ValueError, TypeError):  
                        pass  
                    row_a += 1  
                wb_attr.close()  
            except Exception as e:  
                print(f"[WARN] 读取Attribute失败: {e}")  
  
        # ── 步骤2：读取 AttributeSet 多行格式 ──  
        wb = openpyxl.load_workbook(ATTRSET_XLSX_PATH, read_only=True)  
        ws = wb.worksheets[0]  
        attrsets = []  
        current_set = None  
  
        for row_data in ws.iter_rows(min_row=ATTRSET_DATA_START_ROW, values_only=True):  
            set_id_val  = row_data[1]  # 第2列 (0-indexed=1): SetID  
            attr_id_val = row_data[4]  # 第5列 (0-indexed=4): Attribute.ID  
  
            if set_id_val is not None:  
                # 主行：新的 AttributeSet  
                try:  
                    set_id_int = int(set_id_val)  
                except (ValueError, TypeError):  
                    continue  
                set_name = str(row_data[2] or "")  # 第3列: Name  
                current_set = {  
                    "id": set_id_int,  
                    "name": set_name,  
                    "attrs": []  
                }  
                attrsets.append(current_set)  
  
            # 主行或子行：解析 Attribute ID  
            if attr_id_val is not None and current_set is not None:  
                try:  
                    aid = int(attr_id_val)  
                    current_set["attrs"].append({  
                        "id": aid,  
                        "name": attr_name_map.get(aid, str(aid))  
                    })  
                except (ValueError, TypeError):  
                    pass  
  
        wb.close()  
        return attrsets  
    except Exception as e:  
        print(f"[WARN] 读取AttributeSet失败: {e}")  
        return [] 
  
def read_effects_for_dropdown():  
    """读取Effect数据用于Period效果选择"""  
    try:  
        wb = openpyxl.load_workbook(XLSX_PATH)  
        ws = wb.worksheets[0]  
        effects = []  
        row = DATA_START_ROW  
        while ws.cell(row=row, column=2).value is not None:  
            effect_id = ws.cell(row=row, column=2).value  
            effect_name = ws.cell(row=row, column=3).value  
            if effect_id is not None:  
                effects.append({"id": int(effect_id), "name": str(effect_name or "")})  
            row += 1  
        wb.close()  
        return effects  
    except Exception as e:  
        print(f"[WARN] 读取Effect失败: {e}")  
        return []  
        
# ── Excel 读写层 ─────────────────────────────────────────────────────────────  
  
def init_col_map():  
    """初始化列映射"""  
    global COL_MAP  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.worksheets[0]  
    COL_MAP = {}  
    for col in range(1, 100):  
        header = ws.cell(row=1, column=col).value  
        if header:  
            header = str(header).split('#')[0].strip()  
            if header:  
                COL_MAP[header] = col  
    wb.close()  
    print(f"[INFO] 列映射: {COL_MAP}")  
  
def get_cell_value(ws, row, col_name, default=None):  
    col = COL_MAP.get(col_name)  
    if col is None:  
        return default  
    val = ws.cell(row=row, column=col).value  
    return val if val is not None else default  
  
def parse_int_list(val):  
    if not val or str(val).strip() == '':  
        return []  
    try:  
        return [int(x) for x in re.findall(r"-?\d+", str(val)) if int(x) > 0]  
    except:  
        return []  

def parse_tag_requirement(val, mode):  
    if not val or str(val).strip() == '':  
        return []  
    raw = str(val).strip()  
    parts = raw.split(';')  
    # Requirement format: all;any;none (each section uses comma)
    if len(parts) == 3:  
        if mode == 'all':  
            target = parts[0]  
        elif mode == 'any':  
            target = parts[1]  
        else:  
            target = parts[2]  
        if target == '0' or target.strip() == '':  
            return []  
        return parse_int_list(target)  
    # legacy list format: a;b;c
    return parse_int_list(raw)  

def encode_tag_requirement(vals, mode):  
    nums = [int(x) for x in (vals or []) if int(x) > 0]  
    if not nums:  
        return None  
    tag_csv = ','.join(map(str, nums))  
    if mode == 'all':  
        return f"{tag_csv};0;0"  
    if mode == 'any':  
        return f"0;{tag_csv};0"  
    return f"0;0;{tag_csv}"  
  
def read_effects():  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.worksheets[0]  
    effects = []  
    row = DATA_START_ROW  
  
    while ws.cell(row=row, column=2).value is not None:  
        effect_id = int(ws.cell(row=row, column=2).value)  
        effect = {  
            "id": effect_id,  
            "name": str(get_cell_value(ws, row, "Name", "")),  
            "desc": str(get_cell_value(ws, row, "Desc", "")),  
            "components": [],  
            "assetTags": parse_int_list(get_cell_value(ws, row, "AssetTags")),  
            "grantedTags": parse_int_list(get_cell_value(ws, row, "GrantedTags")),  
            "duration": None,  
            "period": None,  
            "modifiers": [],  
            "cueOnApply": parse_int_list(get_cell_value(ws, row, "CueOnApply")),  
            "cueOnTick": parse_int_list(get_cell_value(ws, row, "CueOnTick")),  
            "cueOnAdd": parse_int_list(get_cell_value(ws, row, "CueOnAdd")),  
            "cueOnRemove": parse_int_list(get_cell_value(ws, row, "CueOnRemove")),  
            "cueOnActivate": parse_int_list(get_cell_value(ws, row, "CueOnActivate")),  
            "cueOnDeactivate": parse_int_list(get_cell_value(ws, row, "CueOnDeactivate")),  
            "grantedAbilities": [],  
            "stacking": None  
        }  

        for field in TAG_REQUIREMENT_PROTOCOL_FIELDS:
            effect[field["json_key"]] = parse_tag_requirement(
                get_cell_value(ws, row, field["excel_header"]), field["mode"]
            )
  
        if effect["assetTags"]: effect["components"].append("AssetTags")  
        if effect["grantedTags"]: effect["components"].append("GrantedTags")  
        for field in TAG_REQUIREMENT_PROTOCOL_FIELDS:
            if effect[field["json_key"]]:
                effect["components"].append(field["component"])
        if effect["cueOnApply"]: effect["components"].append("CueOnApply")  
        if effect["cueOnTick"]: effect["components"].append("CueOnTick")  
        if effect["cueOnAdd"]: effect["components"].append("CueOnAdd")  
        if effect["cueOnRemove"]: effect["components"].append("CueOnRemove")  
        if effect["cueOnActivate"]: effect["components"].append("CueOnActivate")  
        if effect["cueOnDeactivate"]: effect["components"].append("CueOnDeactivate")  
  
        # Duration — 修复：duration_time != 0 以支持 -1 = Infinite  
        duration_col = COL_MAP.get("Duration")  
        if duration_col:  
            duration_unit = ws.cell(row=row, column=duration_col).value  
            if duration_unit is not None:  
                try:  
                    duration_time = int(ws.cell(row=row, column=duration_col + 1).value or 0)  
                    duration_reset = ws.cell(row=row, column=duration_col + 2).value  
                    effect["duration"] = {  
                        "unit": ENUM_TIME_UNIT[int(duration_unit)] if int(duration_unit) < len(ENUM_TIME_UNIT) else "Frame",  
                        "time": duration_time,  
                        "resetStartTimeWhenActivated": str(duration_reset).lower() == "true" if duration_reset else False  
                    }  
                    if duration_time != 0:   # ← 修复：原为 > 0，现改为 != 0 以支持 -1=Infinite  
                        effect["components"].append("Duration")  
                except:  
                    pass  
  
        # Period  
        period_col = COL_MAP.get("Period")  
        if period_col:  
            period_time = ws.cell(row=row, column=period_col).value  
            if period_time is not None and int(period_time) > 0:  
                try:  
                    period_effects = parse_int_list(ws.cell(row=row, column=period_col + 1).value)  
                    period_first = ws.cell(row=row, column=period_col + 2).value  
                    effect["period"] = {  
                        "time": int(period_time),  
                        "effects": period_effects,  
                        "firstTrigger": str(period_first).lower() == "true" if period_first else False  
                    }  
                    effect["components"].append("Period")  
                except:  
                    pass  
  
        # Modifiers  
        modifiers_raw = get_cell_value(ws, row, "Modifiers")  
        if modifiers_raw and str(modifiers_raw).strip():  
            try:  
                for mod_str in str(modifiers_raw).split('|'):  
                    if not mod_str.strip():  
                        continue  
                    parts = mod_str.strip().split(';')  
                    if len(parts) >= 5:  
                        effect["modifiers"].append({  
                            "attrSet": int(parts[0]),  
                            "attr": int(parts[1]),  
                            "magnitude": float(parts[2]),  
                            "operation": ENUM_OPERATION_REV.get(int(parts[3]), "Add"),  
                            "mmc": int(parts[4])  
                        })  
                if effect["modifiers"]:  
                    effect["components"].append("Modifiers")  
            except Exception as e:  
                print(f"[WARN] 解析Modifiers失败: {e}")  
  
        # GrantedAbility  
        ability_raw = get_cell_value(ws, row, "GrantedAbility")  
        if ability_raw and str(ability_raw).strip():  
            try:  
                for ab_str in str(ability_raw).split('|'):  
                    if not ab_str.strip():  
                        continue  
                    parts = ab_str.strip().split(';')  
                    if len(parts) >= 5:  
                        effect["grantedAbilities"].append({  
                            "abilityId": int(parts[0]),  
                            "level": int(parts[1]),  
                            "activationPolicy": ENUM_ABILITY_ACTIVATE[int(parts[2])] if int(parts[2]) < len(ENUM_ABILITY_ACTIVATE) else "None",  
                            "deactivationPolicy": ENUM_ABILITY_DEACTIVATE[int(parts[3])] if int(parts[3]) < len(ENUM_ABILITY_DEACTIVATE) else "None",  
                            "removePolicy": ENUM_ABILITY_REMOVE[int(parts[4])] if int(parts[4]) < len(ENUM_ABILITY_REMOVE) else "None"  
                        })  
                if effect["grantedAbilities"]:  
                    effect["components"].append("GrantedAbility")  
            except Exception as e:  
                print(f"[WARN] 解析GrantedAbility失败: {e}")  
  
        # Stacking — 修复：原为 int(stacking_code) > 0，现改为 stacking_code is not None（code=0 合法）  
        stacking_col = COL_MAP.get("Stacking")  
        if stacking_col:  
            stacking_code = ws.cell(row=row, column=stacking_col).value  
            if stacking_code is not None:   # ← 修复  
                try:  
                    effect["stacking"] = {  
                        "code": int(stacking_code),  
                        "stackingType": ENUM_STACKING_TYPE[int(ws.cell(row=row, column=stacking_col + 1).value or 0)],  
                        "limitCount": int(ws.cell(row=row, column=stacking_col + 2).value or 0),  
                        "durationRefreshPolicy": ENUM_DURATION_REFRESH[int(ws.cell(row=row, column=stacking_col + 3).value or 0)],  
                        "periodResetPolicy": ENUM_PERIOD_RESET[int(ws.cell(row=row, column=stacking_col + 4).value or 0)],  
                        "expirationPolicy": ENUM_EXPIRATION[int(ws.cell(row=row, column=stacking_col + 5).value or 0)],  
                        "denyOverflowApplication": str(ws.cell(row=row, column=stacking_col + 6).value).lower() == "true",  
                        "clearStackOnOverflow": str(ws.cell(row=row, column=stacking_col + 7).value).lower() == "true",  
                        "overflowEffects": parse_int_list(ws.cell(row=row, column=stacking_col + 8).value)  
                    }  
                    effect["components"].append("Stacking")  
                except:  
                    pass  
  
        effects.append(effect)  
        row += 1  
  
    wb.close()  
    return effects 

def write_effects(effects):  
    wb = openpyxl.load_workbook(XLSX_PATH)  
    ws = wb.worksheets[0]  
  
    # 清空旧数据  
    for r in range(DATA_START_ROW, ws.max_row + 1):  
        for c in range(1, 100):  
            ws.cell(row=r, column=c).value = None  
  
    # 写入新数据  
    for i, effect in enumerate(sorted(effects, key=lambda x: x["id"])):  
        row = DATA_START_ROW + i  
        ws.cell(row=row, column=2).value = effect["id"]  
        ws.cell(row=row, column=3).value = effect.get("name", "")  
        ws.cell(row=row, column=4).value = effect.get("desc", "")  
  
        comp = effect.get("components", [])  
  
        # Tags  
        ws.cell(row=row, column=COL_MAP.get("AssetTags", 5)).value = ";".join(map(str, effect.get("assetTags", []))) if "AssetTags" in comp else None  
        ws.cell(row=row, column=COL_MAP.get("GrantedTags", 6)).value = ";".join(map(str, effect.get("grantedTags", []))) if "GrantedTags" in comp else None  
        fallback_col_map = {
            "ApplicationRequiredTags": 7,
            "OngoingRequiredTags": 8,
            "RemoveGameplayEffectsWithTags": 9,
            "ImmunityTags": 10,
        }
        for field in TAG_REQUIREMENT_PROTOCOL_FIELDS:
            ws.cell(row=row, column=COL_MAP.get(field["excel_header"], fallback_col_map[field["excel_header"]])).value = (
                encode_tag_requirement(effect.get(field["json_key"], []), field["mode"])
                if field["component"] in comp else None
            )
  
        # Cues  
        ws.cell(row=row, column=COL_MAP.get("CueOnApply", 11)).value = ";".join(map(str, effect.get("cueOnApply", []))) if "CueOnApply" in comp else None  
        ws.cell(row=row, column=COL_MAP.get("CueOnTick", 12)).value = ";".join(map(str, effect.get("cueOnTick", []))) if "CueOnTick" in comp else None  
        ws.cell(row=row, column=COL_MAP.get("CueOnAdd", 13)).value = ";".join(map(str, effect.get("cueOnAdd", []))) if "CueOnAdd" in comp else None  
        ws.cell(row=row, column=COL_MAP.get("CueOnRemove", 14)).value = ";".join(map(str, effect.get("cueOnRemove", []))) if "CueOnRemove" in comp else None  
        ws.cell(row=row, column=COL_MAP.get("CueOnActivate", 15)).value = ";".join(map(str, effect.get("cueOnActivate", []))) if "CueOnActivate" in comp else None  
        ws.cell(row=row, column=COL_MAP.get("CueOnDeactivate", 16)).value = ";".join(map(str, effect.get("cueOnDeactivate", []))) if "CueOnDeactivate" in comp else None  
  
        # Duration  
        duration_col = COL_MAP.get("Duration")  
        if duration_col and "Duration" in comp:  
            duration = effect.get("duration", {})  
            ws.cell(row=row, column=duration_col).value = ENUM_TIME_UNIT.index(duration.get("unit", "Frame"))  
            ws.cell(row=row, column=duration_col + 1).value = duration.get("time", 0)  
            ws.cell(row=row, column=duration_col + 2).value = duration.get("resetStartTimeWhenActivated", False)  
        elif duration_col:  
            ws.cell(row=row, column=duration_col).value = None  
            ws.cell(row=row, column=duration_col + 1).value = None  
            ws.cell(row=row, column=duration_col + 2).value = None  
  
        # Period  
        period_col = COL_MAP.get("Period")  
        if period_col and "Period" in comp:  
            period = effect.get("period", {})  
            ws.cell(row=row, column=period_col).value = period.get("time", 0)  
            ws.cell(row=row, column=period_col + 1).value = ";".join(map(str, period.get("effects", [])))  
            ws.cell(row=row, column=period_col + 2).value = str(period.get("firstTrigger", False))  
        elif period_col:  
            ws.cell(row=row, column=period_col).value = None  
            ws.cell(row=row, column=period_col + 1).value = None  
            ws.cell(row=row, column=period_col + 2).value = None  
  
        # Modifiers  
        modifiers_col = COL_MAP.get("Modifiers")  
        if modifiers_col and "Modifiers" in comp:  
            mods = effect.get("modifiers", [])  
            mod_strs = []  
            for m in mods:  
                mod_strs.append(  
                    f"{m.get('attrSet', 0)};{m.get('attr', 0)};{m.get('magnitude', 0)};"  
                    f"{ENUM_OPERATION.get(m.get('operation', 'Add'), 0)};{m.get('mmc', 0)}"  
                )  
            ws.cell(row=row, column=modifiers_col).value = "|".join(mod_strs) if mod_strs else None  
        elif modifiers_col:  
            ws.cell(row=row, column=modifiers_col).value = None  
  
        # GrantedAbility  
        ability_col = COL_MAP.get("GrantedAbility")  
        if ability_col and "GrantedAbility" in comp:  
            abilities = effect.get("grantedAbilities", [])  
            ab_strs = []  
            for a in abilities:  
                act_idx   = ENUM_ABILITY_ACTIVATE.index(a.get("activationPolicy", "None"))   if a.get("activationPolicy")   in ENUM_ABILITY_ACTIVATE   else 0  
                deact_idx = ENUM_ABILITY_DEACTIVATE.index(a.get("deactivationPolicy", "None")) if a.get("deactivationPolicy") in ENUM_ABILITY_DEACTIVATE else 0  
                rem_idx   = ENUM_ABILITY_REMOVE.index(a.get("removePolicy", "None"))           if a.get("removePolicy")       in ENUM_ABILITY_REMOVE     else 0  
                ab_strs.append(f"{a.get('abilityId', 0)};{a.get('level', 1)};{act_idx};{deact_idx};{rem_idx}")  
            ws.cell(row=row, column=ability_col).value = "|".join(ab_strs) if ab_strs else None  
        elif ability_col:  
            ws.cell(row=row, column=ability_col).value = None  
  
        # Stacking — 修复：改为检查 "Stacking" in comp，不再用 code > 0 判断  
        stacking_col = COL_MAP.get("Stacking")  
        if stacking_col and "Stacking" in comp:  
            stacking = effect.get("stacking", {})  
            ws.cell(row=row, column=stacking_col).value = stacking.get("code", 0)  
            ws.cell(row=row, column=stacking_col + 1).value = ENUM_STACKING_TYPE.index(stacking.get("stackingType", "AggregateBySource")) if stacking.get("stackingType") in ENUM_STACKING_TYPE else 0  
            ws.cell(row=row, column=stacking_col + 2).value = stacking.get("limitCount", 0)  
            ws.cell(row=row, column=stacking_col + 3).value = ENUM_DURATION_REFRESH.index(stacking.get("durationRefreshPolicy", "NeverRefresh")) if stacking.get("durationRefreshPolicy") in ENUM_DURATION_REFRESH else 0  
            ws.cell(row=row, column=stacking_col + 4).value = ENUM_PERIOD_RESET.index(stacking.get("periodResetPolicy", "NeverRefresh")) if stacking.get("periodResetPolicy") in ENUM_PERIOD_RESET else 0  
            ws.cell(row=row, column=stacking_col + 5).value = ENUM_EXPIRATION.index(stacking.get("expirationPolicy", "ClearEntireStack")) if stacking.get("expirationPolicy") in ENUM_EXPIRATION else 0  
            ws.cell(row=row, column=stacking_col + 6).value = str(stacking.get("denyOverflowApplication", False))  
            ws.cell(row=row, column=stacking_col + 7).value = str(stacking.get("clearStackOnOverflow", False))  
            ws.cell(row=row, column=stacking_col + 8).value = ";".join(map(str, stacking.get("overflowEffects", [])))  
        elif stacking_col:  
            for offset in range(9):  
                ws.cell(row=row, column=stacking_col + offset).value = None  
  
    wb.save(XLSX_PATH)  
    wb.close()

def next_id(effects):  
    return max((e["id"] for e in effects), default=1000) + 1  
  
def validate_effect(effect):  
    """校验Effect数据"""  
    if not effect.get("name"):  
        return "Effect名称不能为空"  
    return None  
  
# ── HTTP 处理器 ──────────────────────────────────────────────────────────────  
  
class Handler(BaseHTTPRequestHandler):  
  
    def log_message(self, fmt, *args):  
        print(f"  [{args[1] if len(args) > 1 else 'INFO'}] {self.command} {self.path}")  
  
    def send_json(self, data, status=200):  
        body = json.dumps(data, ensure_ascii=False).encode()  
        self.send_response(status)  
        self.send_header("Content-Type", "application/json; charset=utf-8")  
        self.send_header("Content-Length", len(body))  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.end_headers()  
        self.wfile.write(body)  
  
    def read_json(self):  
        n = int(self.headers.get("Content-Length", 0))  
        return json.loads(self.rfile.read(n))  
  
    def do_OPTIONS(self):  
        self.send_response(204)  
        self.send_header("Access-Control-Allow-Origin", "*")  
        self.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,DELETE,OPTIONS")  
        self.send_header("Access-Control-Allow-Headers", "Content-Type")  
        self.end_headers()  
  
    def do_GET(self):  
        p = urlparse(self.path).path  
  
        if p == "/":  
            p = "/effect_editor.html"  
  
        # 静态文件服务  
        ext = os.path.splitext(p)[1]  
        if ext in STATIC_TYPES:  
            filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), p.lstrip("/"))  
            if os.path.exists(filepath):  
                body = open(filepath, "rb").read()  
                self.send_response(200)  
                self.send_header("Content-Type", STATIC_TYPES[ext])  
                self.send_header("Content-Length", len(body))  
                self.end_headers()  
                self.wfile.write(body)  
            else:  
                self.send_json({"ok": False, "error": f"File not found: {p}"}, 404)  
            return  
  
        # API 路由  
        if p == "/api/effects":  
            try:  
                self.send_json({"ok": True, "effects": read_effects()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
  
        elif p == "/api/info":  
            self.send_json({"ok": True, "xlsx": XLSX_PATH})  
  
        # ── 修复：将 /api/dropdowns 拆分为 /api/choices/* 子路由 ──────────  
        elif p == "/api/choices/tags":  
            try:  
                self.send_json({"ok": True, "tags": read_tags_for_dropdown()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
  
        elif p == "/api/choices/cues":  
            try:  
                self.send_json({"ok": True, "cues": read_cues_for_dropdown()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
  
        elif p == "/api/choices/abilities":  
            try:  
                self.send_json({"ok": True, "abilities": read_abilities_for_dropdown()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
  
        elif p == "/api/choices/attrsets":  
            try:  
                self.send_json({"ok": True, "attrsets": read_attrsets_for_dropdown()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
  
        elif p == "/api/choices/effects":  
            try:  
                self.send_json({"ok": True, "effects": read_effects_for_dropdown()})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
                
        elif p == "/api/choices/mmcs":  
                    try:  
                        self.send_json({"ok": True, "mmcs": read_mmcs_for_dropdown()})  
                    except Exception as e:  
                        self.send_json({"ok": False, "error": str(e)}, 500)
                        
        elif p == "/api/enums":  
            self.send_json({  
                "ok": True,  
                "timeUnit": ENUM_TIME_UNIT,  
                "operation": list(ENUM_OPERATION.keys()),  
                "stackingType": ENUM_STACKING_TYPE,  
                "durationRefresh": ENUM_DURATION_REFRESH,  
                "periodReset": ENUM_PERIOD_RESET,  
                "expiration": ENUM_EXPIRATION,  
                "abilityActivate": ENUM_ABILITY_ACTIVATE,  
                "abilityDeactivate": ENUM_ABILITY_DEACTIVATE,  
                "abilityRemove": ENUM_ABILITY_REMOVE  
            })  
  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_POST(self):  
        p = urlparse(self.path).path  
        if p == "/api/effects":  
            try:  
                body = self.read_json()  
                effects = read_effects()  
                custom_id = body.get("id")  
                if custom_id is not None:  
                    custom_id = int(custom_id)  
                    if any(e["id"] == custom_id for e in effects):  
                        return self.send_json({"ok": False, "error": f"ID {custom_id} 已存在"}, 400)  
                    new_id = custom_id  
                else:  
                    new_id = next_id(effects)  
  
                new_effect = {  
                    "id": new_id,  
                    "name": body.get("name", "").strip(),  
                    "desc": body.get("desc", "").strip(),  
                    "components": body.get("components", []),  
                    "assetTags": body.get("assetTags", []),  
                    "grantedTags": body.get("grantedTags", []),  
                    "duration": body.get("duration"),  
                    "period": body.get("period"),  
                    "modifiers": body.get("modifiers", []),  
                    "cueOnApply": body.get("cueOnApply", []),  
                    "cueOnTick": body.get("cueOnTick", []),  
                    "cueOnAdd": body.get("cueOnAdd", []),  
                    "cueOnRemove": body.get("cueOnRemove", []),  
                    "cueOnActivate": body.get("cueOnActivate", []),  
                    "cueOnDeactivate": body.get("cueOnDeactivate", []),  
                    "grantedAbilities": body.get("grantedAbilities", []),  
                    "stacking": body.get("stacking")  
                }  

                for field in TAG_REQUIREMENT_PROTOCOL_FIELDS:
                    new_effect[field["json_key"]] = body.get(field["json_key"], [])
  
                err = validate_effect(new_effect)  
                if err:  
                    return self.send_json({"ok": False, "error": err}, 400)  
  
                write_effects(effects + [new_effect])  
                self.send_json({"ok": True, "effect": new_effect})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_PUT(self):  
        parts = urlparse(self.path).path.strip("/").split("/")  
        if len(parts) == 3 and parts[:2] == ["api", "effects"]:  
            try:  
                eid = int(parts[2])  
                body = self.read_json()  
                effects = read_effects()  
                effect = next((e for e in effects if e["id"] == eid), None)  
                if not effect:  
                    return self.send_json({"ok": False, "error": f"ID不存在: {eid}"}, 404)  
  
                new_id = body.get("id")  
                if new_id is not None:  
                    new_id = int(new_id)  
                    if new_id != eid and any(e["id"] == new_id for e in effects):  
                        return self.send_json({"ok": False, "error": f"ID {new_id} 已存在"}, 400)  
                    effect["id"] = new_id  
  
                effect["name"] = body.get("name", effect["name"]).strip()  
                effect["desc"] = body.get("desc", effect.get("desc", "")).strip()  
                effect["components"] = body.get("components", effect["components"])  
                effect["assetTags"] = body.get("assetTags", effect.get("assetTags", []))  
                effect["grantedTags"] = body.get("grantedTags", effect.get("grantedTags", []))  
                for field in TAG_REQUIREMENT_PROTOCOL_FIELDS:
                    effect[field["json_key"]] = body.get(field["json_key"], effect.get(field["json_key"], []))
                effect["duration"] = body.get("duration")  
                effect["period"] = body.get("period")  
                effect["modifiers"] = body.get("modifiers", [])  
                effect["cueOnApply"] = body.get("cueOnApply", effect.get("cueOnApply", []))  
                effect["cueOnTick"] = body.get("cueOnTick", effect.get("cueOnTick", []))  
                effect["cueOnAdd"] = body.get("cueOnAdd", effect.get("cueOnAdd", []))  
                effect["cueOnRemove"] = body.get("cueOnRemove", effect.get("cueOnRemove", []))  
                effect["cueOnActivate"] = body.get("cueOnActivate", effect.get("cueOnActivate", []))  
                effect["cueOnDeactivate"] = body.get("cueOnDeactivate", effect.get("cueOnDeactivate", []))  
                effect["grantedAbilities"] = body.get("grantedAbilities", effect.get("grantedAbilities", []))  
                effect["stacking"] = body.get("stacking")  
  
                err = validate_effect(effect)  
                if err:  
                    return self.send_json({"ok": False, "error": err}, 400)  
  
                write_effects(effects)  
                self.send_json({"ok": True, "effect": effect})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404)  
  
    def do_DELETE(self):  
        parts = urlparse(self.path).path.strip("/").split("/")  
        if len(parts) == 3 and parts[:2] == ["api", "effects"]:  
            try:  
                eid = int(parts[2])  
                effects = read_effects()  
                new_effects = [e for e in effects if e["id"] != eid]  
                if len(new_effects) == len(effects):  
                    return self.send_json({"ok": False, "error": f"ID不存在: {eid}"}, 404)  
                write_effects(new_effects)  
                self.send_json({"ok": True})  
            except Exception as e:  
                self.send_json({"ok": False, "error": str(e)}, 500)  
        else:  
            self.send_json({"ok": False, "error": "Not Found"}, 404) 


# ── 入口 ────────────────────────────────────────────────────────────────────    
  
def main():    
    global XLSX_PATH, TAG_XLSX_PATH, ATTRSET_XLSX_PATH, ABILITY_XLSX_PATH, CUE_XLSX_PATH, MMC_XLSX_PATH, ATTR_XLSX_PATH   
  
    ap = argparse.ArgumentParser(description="EX-GAS GameplayEffect 网页编辑器服务")    
    ap.add_argument("--xlsx", required=True, help="#exgas.gameplayEffects.xlsx 路径")    
    ap.add_argument("--tag-xlsx", help="#exgas.gameplayTags.xlsx 路径")    
    ap.add_argument("--attr-xlsx", help="#exgas.attribute.xlsx 路径")
    ap.add_argument("--attrset-xlsx", help="#exgas.attributeSets.xlsx 路径")    
    ap.add_argument("--ability-xlsx", help="#exgas.abilities.xlsx 路径")    
    ap.add_argument("--cue-xlsx", help="#exgas.cues.xlsx 路径")    
    ap.add_argument("--mmc-xlsx", help="#exgas.mmc.xlsx 路径")    
    ap.add_argument("--port", type=int, default=8769)    
    ap.add_argument("--no-browser", action="store_true")    
    args = ap.parse_args()    
  
    XLSX_PATH = os.path.abspath(args.xlsx)    
    if not os.path.exists(XLSX_PATH):    
        print(f"[ERROR] 文件不存在: {XLSX_PATH}")    
        sys.exit(1)    
  
    if args.tag_xlsx:    
        TAG_XLSX_PATH = os.path.abspath(args.tag_xlsx)    
    if args.attrset_xlsx:    
        ATTRSET_XLSX_PATH = os.path.abspath(args.attrset_xlsx)    
    if args.ability_xlsx:    
        ABILITY_XLSX_PATH = os.path.abspath(args.ability_xlsx)    
    if args.cue_xlsx:    
        CUE_XLSX_PATH = os.path.abspath(args.cue_xlsx)    
    if args.mmc_xlsx:    
        MMC_XLSX_PATH = os.path.abspath(args.mmc_xlsx)   
    if args.attr_xlsx:  
        ATTR_XLSX_PATH = os.path.abspath(args.attr_xlsx) 
  
    init_col_map()    
  
    url = f"http://127.0.0.1:{args.port}"    
    print(f"[EX-GAS Effect Editor] {url}")    
    print(f"  Excel: {XLSX_PATH}")    
    if MMC_XLSX_PATH:    
        print(f"  MMC:   {MMC_XLSX_PATH}")    
    print(f"  Ctrl+C 停止")    
  
    if not args.no_browser:    
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()    
  
    try:    
        HTTPServer(("127.0.0.1", args.port), Handler).serve_forever()    
    except KeyboardInterrupt:    
        print("\n[EX-GAS Effect Editor] 服务已停止")    
  
if __name__ == "__main__":    
    main()
