#!/usr/bin/env python3  
# -*- coding: utf-8 -*-  
"""  
EX-GAS Web编辑器 - 跨表 ID/Name 选项集合  
对标 C# 的 GasXlsxChoice 类，供各子编辑器的 /api/choices/* 接口使用。  
  
依赖: openpyxl  
放置位置: Assets/GAS/Editor/WebEditor/gas_xlsx_choice.py  
  
各 server.py 使用方式：  
    import sys, os  
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  
    from gas_xlsx_choice import GasXlsxChoice  
  
    # 在 main() 中构造：  
    choices = GasXlsxChoice({  
        "tag":    args.tag_xlsx,  
        "attr":   args.attr_xlsx,  
        "attrset": args.attrset_xlsx,  
        "cue":    args.cue_xlsx,  
        "effect": args.effect_xlsx,  
        "ability": args.ability_xlsx,  
        "mmc":    args.mmc_xlsx,  
    })  
"""  
  
import os  
  
try:  
    import openpyxl  
except ImportError:  
    raise ImportError("[gas_xlsx_choice] 请先运行: pip install openpyxl")  
  
  
# ── 通用读取：从指定行开始，第2列=ID，第3列=Name ─────────────────────────────  
def _read_id_name(xlsx_path: str, data_start_row: int = 4) -> list:  
    """  
    通用读取，返回 [{"id": int, "name": str}, ...] 列表。  
    约定：第2列为ID，第3列为Name。遇到第2列为None时停止。  
    对 AttributeSet 等有子行的表，调用方需自行处理（或使用专用函数）。  
    """  
    if not xlsx_path or not os.path.exists(xlsx_path):  
        return []  
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)  
    ws = wb.worksheets[0]  
    result = []  
    row = data_start_row  
    while True:  
        id_val = ws.cell(row=row, column=2).value  
        if id_val is None:  
            break  
        name_val = ws.cell(row=row, column=3).value  
        try:  
            result.append({"id": int(id_val), "name": str(name_val or "")})  
        except (ValueError, TypeError):  
            pass  
        row += 1  
    wb.close()  
    return result  
  
  
# ── AttributeSet 专用：只读主行（跳过 set_id 为 None 的子行） ─────────────────  
def _read_attrset_id_name(xlsx_path: str) -> list:  
    """  
    AttributeSet 表 DATA_START_ROW=5，主行第2列有 set_id，  
    子行第2列为 None（只有 Attribute.ID 在第5列）。  
    这里只返回主行的 {id, name}。  
    """  
    if not xlsx_path or not os.path.exists(xlsx_path):  
        return []  
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)  
    ws = wb.worksheets[0]  
    result = []  
    for row_data in ws.iter_rows(min_row=5, values_only=True):  
        set_id_val = row_data[1]   # 第2列(0-indexed=1)  
        if set_id_val is None:  
            continue               # 跳过子行  
        name_val = row_data[2]    # 第3列(0-indexed=2)  
        try:  
            result.append({"id": int(set_id_val), "name": str(name_val or "")})  
        except (ValueError, TypeError):  
            pass  
    wb.close()  
    return result  
  
  
# ── AttributeSet 专用：返回指定 attrset 内的 Attribute {id, name} 列表 ────────  
def _read_attrs_of_attrset(attrset_xlsx_path: str, attr_xlsx_path: str, attrset_id: int) -> list:  
    """  
    读取指定 AttributeSet 内所有 Attribute 的 {id, name}。  
    需要同时读 attr.xlsx 做 id->name 映射。  
    AttributeSet 表中 Attribute.ID 在第5列(COL_ATTR_ID=5, 0-indexed=4)。  
    """  
    # 先构建 attr id -> name 映射  
    all_attrs = {}  
    for a in _read_id_name(attr_xlsx_path, data_start_row=4):  
        all_attrs[a["id"]] = a["name"]  
  
    if not attrset_xlsx_path or not os.path.exists(attrset_xlsx_path):  
        return []  
    wb = openpyxl.load_workbook(attrset_xlsx_path, read_only=True, data_only=True)  
    ws = wb.worksheets[0]  
    result = []  
    current_set_id = None  
    for row_data in ws.iter_rows(min_row=5, values_only=True):  
        set_id_val = row_data[1]   # 第2列  
        attr_id_val = row_data[4]  # 第5列 Attribute.ID  
        if set_id_val is not None:  
            try:  
                current_set_id = int(set_id_val)  
            except (ValueError, TypeError):  
                current_set_id = None  
        if current_set_id == attrset_id and attr_id_val is not None:  
            try:  
                aid = int(attr_id_val)  
                result.append({"id": aid, "name": all_attrs.get(aid, str(aid))})  
            except (ValueError, TypeError):  
                pass  
    wb.close()  
    return result  
  
  
# ════════════════════════════════════════════════════════════════════════════  
class GasXlsxChoice:  
    """  
    统一提供各配置表的 {id, name} 选项列表（对标 C# GasXlsxChoice）。  
  
    xlsx_paths 字典键名：  
        "tag"     -> #exgas.gameplayTags.xlsx  
        "attr"    -> #exgas.attribute.xlsx  
        "attrset" -> #exgas.attributeSet.xlsx  
        "cue"     -> #exgas.gameplayCue.xlsx  
        "effect"  -> #exgas.gameplayEffect.xlsx  
        "ability" -> #exgas.ability.xlsx  
        "mmc"     -> #exgas.mmc.xlsx  
  
    任何 key 对应的路径为空或文件不存在时，对应方法返回空列表（不抛异常）。  
    """  
  
    def __init__(self, xlsx_paths: dict):  
        self._paths = xlsx_paths  
  
    def _path(self, key: str) -> str:  
        p = self._paths.get(key, "")  
        return os.path.abspath(p) if p else ""  
  
    # ── 已验证的3个 ──────────────────────────────────────────────────────────  
  
    def tags(self) -> list:  
        """  
        GameplayTag: DATA_START_ROW=4, COL_ID=2, COL_NAME=3  
        对标 #exgas.gameplayTags.xlsx  
        已通过 GameplayTag/server.py 的 read_tags() 验证。  
        """  
        return _read_id_name(self._path("tag"), data_start_row=4)  
  
    def attrs(self) -> list:  
        """  
        Attribute: DATA_START_ROW=4, COL_ID=2, COL_NAME=3  
        对标 #exgas.attribute.xlsx  
        已通过 Attribute/server.py 的 read_attrs() 验证。  
        """  
        return _read_id_name(self._path("attr"), data_start_row=4)  
  
    def attrsets(self) -> list:  
        """  
        AttributeSet（只返回主行）: DATA_START_ROW=5, COL_SET_ID=2, COL_SET_NAME=3  
        对标 #exgas.attributeSet.xlsx  
        已通过 AttributeSet/server.py 的 read_attrsets() 验证。  
        """  
        return _read_attrset_id_name(self._path("attrset"))  
  
    def attrs_of_attrset(self, attrset_id: int) -> list:  
        """  
        返回指定 AttributeSet 内的 Attribute {id, name} 列表。  
        需要同时有 "attr" 和 "attrset" 路径。  
        """  
        return _read_attrs_of_attrset(  
            self._path("attrset"),  
            self._path("attr"),  
            attrset_id  
        )  
  
    # C# 侧读取逻辑已确认这4个表均从第4行开始、第2列为ID（参见 GASCenterViewCue.cs / GASCenterViewMmc.cs）  
    # 但各表的完整列结构（CueLogic列偏移、effect的复合字段等）需要截图确认后补充  
    # 目前 _read_id_name 只读 ID 和 Name，对 choice 用途已足够  
  
    def cues(self) -> list:  
        """  
        GameplayCue: DATA_START_ROW=4, COL_ID=2, COL_NAME=3  
        对标 #exgas.gameplayCue.xlsx  
        已通过截图确认表头结构，choice 只需 ID/Name，无需解析 RequiredTag/ImmunityTag/CueLogic 等字段。  
        """  
        return _read_id_name(self._path("cue"), data_start_row=4)
  
    def effects(self) -> list:  
        """  
        GameplayEffect: DATA_START_ROW=4, COL_ID=2, COL_NAME=3  
        对标 #exgas.gameplayEffect.xlsx  
        已通过截图确认表头结构（共30+列，含Duration/Period/Modifiers/Stacking等复合字段）。  
        choice 只需 ID/Name，_read_id_name 通用逻辑已足够，无需解析复合字段。  
        """  
        return _read_id_name(self._path("effect"), data_start_row=4)
  
    def abilities(self) -> list:  
        """  
        Ability: DATA_START_ROW=4, COL_ID=2, COL_NAME=3  
        对标 #exgas.ability.xlsx  
        已通过截图确认表头结构（ID/Name/Desc/Cost/CdEffect/Cd/各Tag列/AbilityLogic + 后续50列流式参数）。  
        choice 只需 ID/Name，_read_id_name 通用逻辑已足够，无需解析 AbilityLogic 及参数列。  
        """  
        return _read_id_name(self._path("ability"), data_start_row=4)
  
    def mmcs(self) -> list:  
        """  
        MMC: DATA_START_ROW=4, COL_ID=2, COL_NAME=3  
        对标 #exgas.mmc.xlsx  
        已通过截图确认表头结构（ID/Name/Desc/MmcLogic + 后续流式参数列）。  
        choice 只需 ID/Name，_read_id_name 通用逻辑已足够，无需解析 MmcLogic 及参数列。  
        """  
        return _read_id_name(self._path("mmc"), data_start_row=4)
  
    # ── 聚合接口（供前端 /api/choices 一次性获取所有选项） ─────────────────────  
  
    def all_choices(self) -> dict:  
        """  
        返回所有配置表的 choice 列表，供前端一次性加载。  
        各 server.py 可通过 GET /api/choices 暴露此接口。  
        """  
        return {  
            "tags":     self.tags(),  
            "attrs":    self.attrs(),  
            "attrsets": self.attrsets(),  
            "cues":     self.cues(),  
            "effects":  self.effects(),  
            "abilities": self.abilities(),  
            "mmcs":     self.mmcs(),  
        }